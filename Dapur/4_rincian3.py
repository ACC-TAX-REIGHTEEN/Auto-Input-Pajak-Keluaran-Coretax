import openpyxl
import glob
import sys
from openpyxl.styles import PatternFill, Font, Border, Side

file_list = glob.glob("Laporan_Analisa_Pajak_*.xlsx")

if not file_list:
    print("--> File tidak ditemukan")
    sys.exit()

file_name = file_list[0]
print("--> Membuka file Excel:", file_name)

wb = openpyxl.load_workbook(file_name)

if "Rincian" not in wb.sheetnames:
    print("--> Struktur file tidak valid")
    sys.exit()

ws = wb["Rincian"]

start_row = 0
for r in range(1, ws.max_row + 1):
    cell_val = ws.cell(row=r, column=1).value
    if isinstance(cell_val, str) and cell_val.strip() == "C. DATA MATCHED":
        start_row = r + 1
        break

if start_row == 0:
    print("--> Struktur file tidak valid")
    sys.exit()

print("--> Mengambil data dan menyaring catatan")
headers = []
for c in range(1, 11):
    headers.append(ws.cell(row=start_row, column=c).value)

data_filtered = []
for r in range(start_row + 1, ws.max_row + 1):
    catatan = ws.cell(row=r, column=10).value
    if isinstance(catatan, str) and catatan.strip() == "Nama Beda; Format NPWP Beda":
        row_data = []
        for c in range(1, 11):
            row_data.append(ws.cell(row=r, column=c).value)
        data_filtered.append(row_data)

if len(data_filtered) > 0:
    print("--> Mengatur posisi dan membuat sheet Rincian 3")
    sheets = wb.sheetnames
    
    if "Rincian 3" in sheets:
        del wb["Rincian 3"]

    sheets = wb.sheetnames
    if "Rincian 2" in sheets:
        idx = sheets.index("Rincian 2")
        ws_new = wb.create_sheet("Rincian 3", idx + 1)
    else:
        ws_new = wb.create_sheet("Rincian 3")

    ws_new.append(headers)
    for d in data_filtered:
        ws_new.append(d)

    print("--> Memberikan gaya pada tabel")
    header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws_new.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = bold_font
        cell.border = thin_border

    for row_num in range(2, len(data_filtered) + 2):
        for col_num in range(1, len(headers) + 1):
            cell = ws_new.cell(row=row_num, column=col_num)
            cell.border = thin_border

    print("--> Mengatur auto-fit kolom")
    for col in ws_new.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val_str = str(cell.value)
                if cell.value is not None and len(val_str) > max_length:
                    max_length = len(val_str)
            except:
                pass
        ws_new.column_dimensions[col_letter].width = max_length + 2
else:
    print("--> Tidak ada data untuk ditampilkan pada Rincian 3")

print("--> Menyimpan dan menutup file")
wb.save(file_name)
wb.close()
print("--> Proses selesai")
