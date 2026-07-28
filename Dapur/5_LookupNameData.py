import glob
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd


def rapikan_dan_warnai_sheet(ws, hex_color="1F4E79", font_color="FFFFFF"):
    header_fill = PatternFill(
        start_color=hex_color, end_color=hex_color, fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color=font_color)
    header_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            val_str = str(cell.value) if cell.value is not None else ""
            lines = val_str.split("\n")
            len_line = max(len(l) for l in lines) if lines else 0
            if len_line > max_len:
                max_len = len_line

        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


def fix_scientific_notation(val):
    if val is None or pd.isna(val):
        return ""

    if isinstance(val, (float, int)):
        return f"{val:.0f}"

    val_str = str(val).strip()

    if "E" in val_str.upper():
        try:
            clean_str = val_str.replace(",", ".")
            num = float(clean_str)
            return f"{num:.0f}"
        except ValueError:
            return val_str

    return val_str


def jalankan_analisis_pajak():
    files = glob.glob("Laporan_Analisa_Pajak*.xlsx")

    if not files:
        print(
            "--> File 'Laporan_Analisa_Pajak............xlsx' tidak ditemukan di"
            " folder ini."
        )
        return

    file_pajak = files[0]
    file_cleaner = "Hasil_CleanerACC.xlsx"

    print(f"--> Memproses file utama: {file_pajak}")

    if not os.path.exists(file_cleaner):
        print(f"--> File '{file_cleaner}' tidak ditemukan!")
        return

    df_cleaner = pd.read_excel(file_cleaner)

    lookup_dict = {}
    for _, row in df_cleaner.iterrows():
        pelanggan = (
            str(row["Nama Pelanggan"]).strip()
            if pd.notna(row.get("Nama Pelanggan"))
            else ""
        )
        negara = (
            str(row["Negara"]).strip() if pd.notna(row.get("Negara")) else ""
        )
        if pelanggan:
            lookup_dict[pelanggan] = negara

    wb = openpyxl.load_workbook(file_pajak)

    if "Data Asli Accurate" in wb.sheetnames:
        ws_acc = wb["Data Asli Accurate"]

        col_faktur_acc = None
        for col in range(1, ws_acc.max_column + 1):
            val = ws_acc.cell(row=1, column=col).value
            if val and "No. Faktur Pajak" in str(val):
                col_faktur_acc = col
                break

        if col_faktur_acc:
            for r in range(2, ws_acc.max_row + 1):
                cell = ws_acc.cell(row=r, column=col_faktur_acc)
                if cell.value is not None:
                    cell.value = fix_scientific_notation(cell.value)
                    cell.number_format = "@"

    if "Data Asli Coretax" in wb.sheetnames:
        ws_coretax = wb["Data Asli Coretax"]

        col_faktur_coretax = None
        for col in range(1, ws_coretax.max_column + 1):
            val = ws_coretax.cell(row=1, column=col).value
            if val and "Faktur Pajak" in str(val):
                col_faktur_coretax = col
                break

        if col_faktur_coretax:
            for r in range(2, ws_coretax.max_row + 1):
                cell = ws_coretax.cell(row=r, column=col_faktur_coretax)
                if cell.value is not None:
                    cell.value = fix_scientific_notation(cell.value)
                    cell.number_format = "@"
                    
    if "Data Asli Coretax" in wb.sheetnames:
        ws_coretax = wb["Data Asli Coretax"]

        col_faktur_coretax = None
        for col in range(1, ws_coretax.max_column + 1):
            val = ws_coretax.cell(row=1, column=col).value
            if val and "Kode dan Nomor Seri Faktur Pajak yang Diganti/Diretur" in str(val):
                col_faktur_coretax = col
                break

        if col_faktur_coretax:
            for r in range(2, ws_coretax.max_row + 1):
                cell = ws_coretax.cell(row=r, column=col_faktur_coretax)
                if cell.value is not None:
                    cell.value = fix_scientific_notation(cell.value)
                    cell.number_format = "@"

    if "Rincian 3" in wb.sheetnames:
        ws_r3 = wb["Rincian 3"]

        headers = [
            ws_r3.cell(row=1, column=c).value
            for c in range(1, ws_r3.max_column + 1)
        ]

        col_nama_acc = None
        col_nama_coretax = None

        for idx, h in enumerate(headers, 1):
            if h and "Nama Accurate" in str(h):
                col_nama_acc = idx
            elif h and "Nama Coretax" in str(h):
                col_nama_coretax = idx

        mismatched_rows = []
        new_headers = headers + ["Negara (Hasil Lookup)"]

        for r in range(2, ws_r3.max_row + 1):
            val_acc = (
                ws_r3.cell(row=r, column=col_nama_acc).value
                if col_nama_acc
                else ""
            )
            val_coretax = (
                ws_r3.cell(row=r, column=col_nama_coretax).value
                if col_nama_coretax
                else ""
            )

            str_acc = str(val_acc).strip() if val_acc is not None else ""
            str_coretax = (
                str(val_coretax).strip() if val_coretax is not None else ""
            )

            negara_lookup = lookup_dict.get(str_acc, "TIDAK DITEMUKAN")

            if negara_lookup.upper() != str_coretax.upper():
                row_data = [
                    ws_r3.cell(row=r, column=c).value
                    for c in range(1, ws_r3.max_column + 1)
                ]
                row_data.append(negara_lookup)
                mismatched_rows.append(row_data)

        if "Nama Negara Beda" in wb.sheetnames:
            del wb["Nama Negara Beda"]

        ws_new = wb.create_sheet(title="Nama Negara Beda")
        ws_new.append(new_headers)

        for row in mismatched_rows:
            ws_new.append(row)

    for sheet in wb.worksheets:
        rapikan_dan_warnai_sheet(sheet, hex_color="1F4E79", font_color="FFFFFF")

    wb.save(file_pajak)
    print(
        f"--> Selesai! File '{file_pajak}' berhasil diperbarui, sheet 'Nama"
        " Negara Beda' telah dibuat."
    )


if __name__ == "__main__":
    jalankan_analisis_pajak()