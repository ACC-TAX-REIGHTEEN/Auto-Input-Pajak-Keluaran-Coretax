import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

def process_tax_report():
    file_pattern = "Laporan_Analisa_Pajak_*.xlsx"
    files = glob.glob(file_pattern)

    if not files:
        print("--> File tidak ditemukan.")
        return

    target_file = files[0]
    sheet_name = "Rincian"
    output_sheet = "Rincian 2"

    try:
        df = pd.read_excel(target_file, sheet_name=sheet_name, header=None)
    except Exception as e:
        print(f"--> Error membaca file: {e}")
        return

    def find_row_index(keyword, start_search_from=0):
        subset = df.iloc[start_search_from:, 0].astype(str)
        matches = subset[subset.str.contains(keyword, case=False, na=False)]
        if not matches.empty:
            return matches.index[0]
        return None

    idx_start_a = find_row_index("A. DATA YANG TIDAK ADA DI ACCURATE")
    idx_end_a = find_row_index("Total A", start_search_from=idx_start_a if idx_start_a is not None else 0)
    idx_start_b = find_row_index("B. DATA YANG TIDAK ADA DI CORETAX")
    idx_end_b = find_row_index("Total B", start_search_from=idx_start_b if idx_start_b is not None else 0)

    if None in [idx_start_a, idx_end_a, idx_start_b, idx_end_b]:
        print("--> Struktur file tidak valid.")
        return

    df_a = df.iloc[idx_start_a + 2 : idx_end_a, [1, 3, 5, 7]].copy()
    df_a.columns = ['Tanggal', 'Nama', 'Nomor Pajak', 'Jumlah']
    df_a['Keterangan'] = 'Data Tidak Ada di Accurate'

    df_b = df.iloc[idx_start_b + 2 : idx_end_b, [0, 2, 4, 6]].copy()
    df_b.columns = ['Tanggal', 'Nama', 'Nomor Pajak', 'Jumlah']
    df_b['Keterangan'] = 'Data Tidak Ada Di Coretax'

    df_combined = pd.concat([df_a, df_b], ignore_index=True)
    df_combined = df_combined.dropna(subset=['Tanggal', 'Jumlah'], how='all')
    df_combined['Jumlah'] = pd.to_numeric(df_combined['Jumlah'], errors='coerce').fillna(0)

    month_map = {
        'Jan': 'Jan', 'Feb': 'Feb', 'Mar': 'Mar', 'Apr': 'Apr', 'Mei': 'May', 'Jun': 'Jun',
        'Jul': 'Jul', 'Agu': 'Aug', 'Sep': 'Sep', 'Okt': 'Oct', 'Nop': 'Nov', 'Des': 'Dec'
    }

    def parse_indo_date(date_val):
        if pd.isna(date_val): 
            return pd.NaT
        if isinstance(date_val, pd.Timestamp): 
            return date_val
        date_str = str(date_val)
        for indo, eng in month_map.items():
            if indo in date_str:
                date_str = date_str.replace(indo, eng)
                break
        return pd.to_datetime(date_str, errors='coerce')

    df_combined['Tanggal_Sort'] = df_combined['Tanggal'].apply(parse_indo_date)
    df_sorted = df_combined.sort_values(by=['Tanggal_Sort', 'Jumlah'], ascending=[True, True])
    
    cols_final = ['Tanggal', 'Nama', 'Nomor Pajak', 'Jumlah', 'Keterangan']
    df_final = df_sorted[cols_final]

    print(f"--> Menyimpan ke {target_file} di sheet {output_sheet}...")

    with pd.ExcelWriter(target_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_final.to_excel(writer, sheet_name=output_sheet, index=False)
        
        wb = writer.book
        ws = writer.sheets[output_sheet]
        
        header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
        highlight_fill = PatternFill(start_color="88C810", end_color="88C810", fill_type="solid")
        header_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
            keterangan_val = row[4].value 
            
            for cell in row:
                cell.border = thin_border
                if keterangan_val == "Data Tidak Ada Di Coretax":
                    cell.fill = highlight_fill

        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
                if column_letter == 'D' and cell.row > 1:
                    cell.number_format = '#,##0'
            
            ws.column_dimensions[column_letter].width = max_length + 2

        if "Rincian" in wb.sheetnames:
            sheets = wb._sheets
            rincian_idx = [s.title for s in sheets].index("Rincian")
            rincian_2_sheet = wb[output_sheet]
            sheets.remove(rincian_2_sheet)
            sheets.insert(rincian_idx + 1, rincian_2_sheet)

    print("--> Selesai.")

if __name__ == "__main__":
    process_tax_report()