import glob
import sys
from collections import defaultdict
import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

file_list = glob.glob("*.xlsx")

if not file_list:
    print("--> Error: Tidak ada file Excel (.xlsx) yang ditemukan.")
    sys.exit()

file_name = file_list[0]
print("--> Membuka file Excel:", file_name)

wb = openpyxl.load_workbook(file_name)

target_sheet = "Data Asli Coretax"
if target_sheet not in wb.sheetnames:
    print(f"--> Error: Sheet '{target_sheet}' tidak ditemukan dalam file.")
    sys.exit()

ws = wb[target_sheet]

headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(row=1, column=col).value)

focus_columns = [
    "Name",
    "TIN",
    "DocumentDate",
    "Harga Jual/Penggantian/DPP (Rupiah)",
    "DPP Nilai Lain/ DPP (Rupiah)",
    "VAT"
]

col_indices = {}
for col_name in focus_columns:
    if col_name in headers:
        col_indices[col_name] = headers.index(col_name) + 1
    else:
        print(f"--> Peringatan: Kolom '{col_name}' tidak ditemukan pada header.")

print("--> Membaca data dan menganalisis duplikat...")

grouped_rows = defaultdict(list)

for r in range(2, ws.max_row + 1):
    row_data = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]

    if all(val is None for val in row_data):
        continue

    key = tuple(ws.cell(row=r, column=col_indices[col_name]).value for col_name in focus_columns if col_name in col_indices)
    grouped_rows[key].append(row_data)

duplicate_rows = []
for key, rows in grouped_rows.items():
    if len(rows) > 1:
        duplicate_rows.extend(rows)

print(f"--> Ditemukan {len(duplicate_rows)} baris data duplikat.")

new_sheet_name = "Data Coretax Duplikat"

if new_sheet_name in wb.sheetnames:
    del wb[new_sheet_name]

ws_new = wb.create_sheet(title=new_sheet_name)

ws_new.append(headers)
for row in duplicate_rows:
    ws_new.append(row)

if len(duplicate_rows) > 0:
    print("--> Memberikan gaya dan merapikan tabel...")

    header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws_new.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = bold_font
        cell.border = thin_border

    for row_num in range(2, len(duplicate_rows) + 2):
        for col_num in range(1, len(headers) + 1):
            cell = ws_new.cell(row=row_num, column=col_num)
            cell.border = thin_border

    for col in ws_new.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val_str = str(cell.value) if cell.value is not None else ""
                if len(val_str) > max_length:
                    max_length = len(val_str)
            except:
                pass
        ws_new.column_dimensions[col_letter].width = max(max_length + 3, 10)

wb.save(file_name)
wb.close()
print("--> Proses selesai! Hasil telah disimpan.")