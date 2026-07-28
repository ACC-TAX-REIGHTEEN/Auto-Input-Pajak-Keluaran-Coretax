import glob
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd


def audit_final_fix():
    folder_sumber = "."
    pola_file_sumber = "*.xlsx"
    file_gabungan = "Hasil_merger_all_ex2ex_keluaran.xlsx"
    file_output_audit = "Laporan_audit_merger_all_ex2ex_keluaran.xlsx"

    IDX_KEY = 2
    IDX_VAL = 6

    print("--> --- MEMULAI AUDIT PRESISI (MULTI-VALUE SUPPORT) ---")
    print(f"--> Membaca Master: {file_gabungan}...")
    try:
        df_master = pd.read_excel(file_gabungan, header=0, dtype=str)
    except:
        df_master = pd.read_csv(file_gabungan, header=0, dtype=str)

    master_dict = {}

    count_master = 0
    for index, row in df_master.iterrows():
        try:
            if len(row) > IDX_VAL:
                raw_key = (
                    str(row.iloc[IDX_KEY])
                    .strip()
                    .replace(".", "")
                    .replace("-", "")
                    .replace(" ", "")
                )

                raw_val_str = (
                    str(row.iloc[IDX_VAL]).replace(",", "").replace(" ", "")
                )
                if "(" in raw_val_str and ")" in raw_val_str:
                    raw_val_str = "-" + raw_val_str.replace("(", "").replace(
                        ")", ""
                    )

                try:
                    val_float = float(raw_val_str)
                except:
                    val_float = 0.0

                if raw_key not in master_dict:
                    master_dict[raw_key] = []
                master_dict[raw_key].append(val_float)
                count_master += 1
        except:
            continue

    print(f"--> Data Master dimuat: {count_master} baris data.\n")

    list_files = glob.glob(os.path.join(folder_sumber, pola_file_sumber))
    list_files = [
        f
        for f in list_files
        if file_gabungan not in f and "Laporan_Audit" not in f
    ]

    summary_data = []
    detail_errors = []

    for file in list_files:
        print(f"--> Meng-audit: {os.path.basename(file)} ...")

        try:
            try:
                df_src = pd.read_excel(file, header=0, dtype=str)
            except:
                df_src = pd.read_csv(file, header=0, dtype=str)
            df_src.dropna(how="all", inplace=True)

            src_total = 0
            found_total = 0
            err_count = 0

            for idx, row in df_src.iterrows():
                if len(row) <= IDX_VAL:
                    continue

                key_ori = str(row.iloc[IDX_KEY])
                key_clean = (
                    key_ori.strip()
                    .replace(".", "")
                    .replace("-", "")
                    .replace(" ", "")
                )

                val_raw_str = (
                    str(row.iloc[IDX_VAL]).replace(",", "").replace(" ", "")
                )
                if "(" in val_raw_str and ")" in val_raw_str:
                    val_raw_str = "-" + val_raw_str.replace("(", "").replace(
                        ")", ""
                    )
                try:
                    val_src = float(val_raw_str)
                except:
                    val_src = 0.0

                src_total += val_src

                found = False
                matched_val = 0

                if key_clean in master_dict:
                    possible_values = master_dict[key_clean]

                    best_match_idx = -1
                    min_diff = float("inf")

                    for i, p_val in enumerate(possible_values):
                        diff = abs(val_src - p_val)
                        if diff < 1.0:
                            best_match_idx = i
                            min_diff = diff
                            break

                    if best_match_idx != -1:
                        found = True
                        matched_val = possible_values[best_match_idx]
                    else:
                        matched_val = possible_values[0]
                        found = False

                else:
                    found = False
                    matched_val = 0

                if found:
                    found_total += matched_val
                else:
                    err_count += 1
                    if key_clean in master_dict:
                        pesan = "SELISIH NILAI (ID Ada, Nilai Beda)"
                    else:
                        pesan = "DATA HILANG (ID Tidak Ada)"

                    detail_errors.append(
                        {
                            "Nama File": os.path.basename(file),
                            "Baris": idx + 2,
                            "Nomor Faktur": key_ori,
                            "Masalah": pesan,
                            "Nilai Source": val_src,
                            "Nilai Master (Saran)": matched_val
                            if key_clean in master_dict
                            else 0,
                            "Selisih": val_src - matched_val,
                        }
                    )

            summary_data.append(
                {
                    "Nama File": os.path.basename(file),
                    "Total Baris": len(df_src),
                    "Total PPN Source": src_total,
                    "Total PPN Found": found_total,
                    "Selisih": src_total - found_total,
                    "Status": "AMAN" if err_count == 0 else "CEK DETAIL",
                }
            )

        except Exception as e:
            print(f"--> Error {file}: {e}")

    if summary_data:
        print("--> Menyimpan Laporan Presisi...")
        with pd.ExcelWriter(file_output_audit, engine="openpyxl") as writer:
            pd.DataFrame(summary_data).to_excel(
                writer, sheet_name="Ringkasan", index=False
            )
            if detail_errors:
                pd.DataFrame(detail_errors).to_excel(
                    writer, sheet_name="Detail Kesalahan", index=False
                )
            else:
                pd.DataFrame({"Info": ["Perfect Match"]}).to_excel(
                    writer, sheet_name="Detail Kesalahan", index=False
                )

        wb = load_workbook(file_output_audit)
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                            if isinstance(cell.value, (int, float)):
                                h_val = str(ws[f"{col_letter}1"].value).lower()
                                if any(
                                    x in h_val
                                    for x in ["nilai", "ppn", "selisih", "total"]
                                ):
                                    cell.number_format = "#,##0"
                                else:
                                    cell.number_format = "0"
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_len + 3

        wb.save(file_output_audit)
        print(f"--> SELESAI. File output: {file_output_audit}")
    else:
        print("--> Tidak ada data.")


if __name__ == "__main__":
    audit_final_fix()