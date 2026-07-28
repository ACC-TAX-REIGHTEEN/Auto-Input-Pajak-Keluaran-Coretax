import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def gabung_paksa_urutan():
    path_folder = '.' 
    pola_file = "data_export*.xlsx"
    nama_sheet_target = "data" 
    nama_file_output = "HMA_ex2ex.xlsx"

    semua_file = glob.glob(os.path.join(path_folder, pola_file))
    semua_file = [f for f in semua_file if nama_file_output not in f]

    if not semua_file:
        print(f"--> Tidak ditemukan file dengan pola '{pola_file}'.")
        return

    print(f"--> Ditemukan {len(semua_file)} file. Memulai penggabungan...")
    
    list_data = []
    file_sukses = []
    header_utama = None 
    
    for urutan, file in enumerate(semua_file):
        try:
            df = pd.read_excel(file, sheet_name=nama_sheet_target, header=0, dtype=str)
            df.dropna(how='all', inplace=True) 
            
            if header_utama is None:
                header_utama = df.columns.tolist()
            
            df.columns = range(df.shape[1])
            
            if len(df.columns) != len(header_utama):
                print(f"--> [WARNING] {file} jumlah kolom beda. Tetap digabung.")
            
            list_data.append(df)
            file_sukses.append(file)
            print(f"--> [OK] {file} : {len(df)} baris")
            
        except Exception as e:
            print(f"--> [SKIP] {file}: {e}")

    if list_data:
        try:
            print("--> Menggabungkan data...")
            df_gabungan = pd.concat(list_data, axis=0, ignore_index=True)
            
            jumlah_kolom_final = df_gabungan.shape[1]
            if len(header_utama) == jumlah_kolom_final:
                df_gabungan.columns = header_utama
            else:
                print(f"--> Info: Jumlah kolom final ({jumlah_kolom_final}) berbeda dengan header awal.")
                df_gabungan.columns = header_utama + [f"Extra_{i}" for i in range(jumlah_kolom_final - len(header_utama))]

            for col in df_gabungan.columns:
                nama_col_kecil = str(col).lower()
                if any(x in nama_col_kecil for x in ['ppn', 'harga', 'dpp', 'nilai', 'total']):
                    df_gabungan[col] = pd.to_numeric(df_gabungan[col], errors='coerce')

            print(f"--> Menyimpan ke {nama_file_output}...")
            df_gabungan.to_excel(nama_file_output, index=False)
            
            print("--> Merapikan tampilan (Auto-fit)...")
            wb = load_workbook(nama_file_output)
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
            
            for column in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            item_len = len(str(cell.value))
                            if item_len > max_length: 
                                max_length = item_len
                    except: 
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(nama_file_output)
            print(f"--> SUKSES! Data tersimpan di: {nama_file_output}")

            print("--> --- MEMBERSIHKAN FILE ASLI ---")
            for f in file_sukses:
                try:
                    if os.path.exists(f):
                        os.remove(f)
                        print(f"--> [DELETED] {f}")
                except Exception as e:
                    print(f"--> [GAGAL HAPUS] {f}: {e}")
            print("--> Selesai membersihkan file sumber.")

        except Exception as e:
            print(f"--> FATAL ERROR saat menyimpan: {e}")
            print("--> File asli TIDAK dihapus karena proses gagal.")
            
    else:
        print("--> Gagal. Tidak ada data yang berhasil diproses.")

if __name__ == "__main__":
    gabung_paksa_urutan()