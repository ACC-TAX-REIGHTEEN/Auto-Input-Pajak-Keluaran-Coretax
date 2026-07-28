import datetime
import glob
import io
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd


def print_debug_csv(message):
    waktu = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"--> [CSV-LOG] [{waktu}] {message}")


def clean_and_parse_csv_smart(file_path):
    cleaned_rows = []
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()
        if len(lines) == 0:
            return pd.DataFrame()
        header_raw = lines[0].strip()
        header_cols = header_raw.split(",")
        target_col_count = len(header_cols)
        cleaned_rows.append("|".join(header_cols))
        for line in lines[1:]:
            line = line.strip()
            if not line:
                continue
            parts = line.split(",")
            if len(parts) == target_col_count:
                cleaned_rows.append("|".join(parts))
            elif len(parts) > target_col_count:
                cols_right_count = target_col_count - 2
                col_index = parts[0]
                cols_right = parts[-cols_right_count:]
                name_parts = parts[1:-cols_right_count]
                col_name_fixed = " ".join(name_parts).replace('"', "")
                new_row = [col_index, col_name_fixed] + cols_right
                cleaned_rows.append("|".join(new_row))
            else:
                cleaned_rows.append("|".join(parts))
    full_str = "\n".join(cleaned_rows)
    return pd.read_csv(io.StringIO(full_str), sep="|", dtype=str)


def jalankan_proses_csv():
    output_excel = "HMA_csv2ex.xlsx"
    try:
        print_debug_csv("=== MEMULAI PROSES CSV ===")
        pattern = "data_export*.csv"
        all_files = glob.glob(pattern)
        if not all_files:
            print_debug_csv(
                f"Info: Tidak ditemukan file CSV dengan pola '{pattern}'. Skip proses CSV."
            )
            return
        print_debug_csv(
            f"Ditemukan {len(all_files)} file CSV untuk diproses."
        )
        df_list = []
        processed_files = []
        for filename in all_files:
            try:
                df_parsed = clean_and_parse_csv_smart(filename)
                df_list.append(df_parsed)
                processed_files.append(filename)
            except Exception as e:
                print_debug_csv(f"   [ERROR] Gagal memproses {filename}: {e}")
        if not df_list:
            print_debug_csv(
                "Tidak ada data yang berhasil diproses dari file CSV."
            )
            return
        print_debug_csv("Menggabungkan data...")
        df_combined = pd.concat(df_list, ignore_index=True)
        if "Unnamed: 0" in df_combined.columns:
            df_combined = df_combined.drop(columns=["Unnamed: 0"])
        elif len(df_combined.columns) > 0:
            if (
                df_combined.columns[0] == ""
                or "Unnamed" in df_combined.columns[0]
            ):
                df_combined = df_combined.drop(columns=[df_combined.columns[0]])
        col_names = df_combined.columns
        start_numeric_index = 4
        if "TaxBase" in col_names:
            start_numeric_index = df_combined.columns.get_loc("TaxBase")
            print_debug_csv(
                f"   Deteksi Kolom Angka (TaxBase) di kolom index: {start_numeric_index}"
            )
        else:
            print_debug_csv(
                f"   Menggunakan default kolom angka mulai index: {start_numeric_index}"
            )
        for i in range(start_numeric_index, len(col_names)):
            col = col_names[i]
            df_combined[col] = pd.to_numeric(df_combined[col], errors="coerce")
        print_debug_csv(f"Menyimpan ke {output_excel}...")
        with pd.ExcelWriter(output_excel, engine="xlsxwriter") as writer:
            df_combined.to_excel(writer, index=False, sheet_name="Data")
            workbook = writer.book
            worksheet = writer.sheets["Data"]
            fmt_text = workbook.add_format({"num_format": "@"})
            fmt_num = workbook.add_format({"num_format": "#,##0"})
            for i, col in enumerate(df_combined.columns):
                try:
                    max_len = (
                        max(
                            df_combined[col].astype(str).map(len).max(),
                            len(str(col)),
                        )
                        + 2
                    )
                except:
                    max_len = 15
                if i < start_numeric_index:
                    worksheet.set_column(i, i, max_len, fmt_text)
                else:
                    worksheet.set_column(i, i, max_len, fmt_num)
        print_debug_csv(f"SUKSES! File tersimpan: {output_excel}")
        print_debug_csv("--- MEMBERSIHKAN FILE CSV ASLI ---")
        deleted_count = 0
        for csv_file in processed_files:
            try:
                if os.path.exists(csv_file):
                    os.remove(csv_file)
                    print_debug_csv(f"   [DELETED] {csv_file}")
                    deleted_count += 1
            except Exception as e:
                print_debug_csv(f"   [GAGAL HAPUS] {csv_file}: {e}")
        print_debug_csv(f"Selesai. {deleted_count} file CSV telah dihapus.")
    except Exception as e:
        print(f"--> FATAL ERROR CSV: {e}")
        print("--> File CSV TIDAK dihapus karena terjadi error.")


def jalankan_proses_excel():
    path_folder = "."
    pola_file = "data_export*.xlsx"
    nama_sheet_target = "data"
    nama_file_output = "HMA_ex2ex.xlsx"
    print("--> [XLSX-LOG] Memulai Cek File Excel...")
    semua_file = glob.glob(os.path.join(path_folder, pola_file))
    semua_file = [f for f in semua_file if nama_file_output not in f]
    if not semua_file:
        print(
            f"--> [XLSX-LOG] Info: Tidak ditemukan file XLSX dengan pola '{pola_file}'. Skip proses Excel."
        )
        return
    print(
        f"--> Ditemukan {len(semua_file)} file XLSX. Memulai penggabungan..."
    )
    list_data = []
    file_sukses = []
    header_utama = None
    for urutan, file in enumerate(semua_file):
        try:
            df = pd.read_excel(
                file, sheet_name=nama_sheet_target, header=0, dtype=str
            )
            df.dropna(how="all", inplace=True)
            if header_utama is None:
                header_utama = df.columns.tolist()
            df.columns = range(df.shape[1])
            if len(df.columns) != len(header_utama):
                print(
                    f"--> [WARNING] {file} jumlah kolom beda. Tetap digabung."
                )
            list_data.append(df)
            file_sukses.append(file)
            print(f"--> [OK] {file} : {len(df)} baris")
        except Exception as e:
            print(f"--> [SKIP] {file}: {e}")
    if list_data:
        try:
            print("--> Menggabungkan data...")
            df_gabungan = pd.concat(list_data, axis=0, ignore_index=True)
            jumlah_kolom_final = df_gabungan.shape[1]
            if len(header_utama) == jumlah_kolom_final:
                df_gabungan.columns = header_utama
            else:
                print(
                    f"--> Info: Jumlah kolom final ({jumlah_kolom_final}) berbeda dengan header awal."
                )
                df_gabungan.columns = header_utama + [
                    f"Extra_{i}"
                    for i in range(
                        jumlah_kolom_final - len(header_utama)
                    )
                ]
            for col in df_gabungan.columns:
                nama_col_kecil = str(col).lower()
                if any(
                    x in nama_col_kecil
                    for x in ["ppn", "harga", "dpp", "nilai", "total"]
                ):
                    df_gabungan[col] = pd.to_numeric(
                        df_gabungan[col], errors="coerce"
                    )
            print(f"--> Menyimpan ke {nama_file_output}...")
            df_gabungan.to_excel(nama_file_output, index=False)
            print("--> Merapikan tampilan (Auto-fit)...")
            wb = load_workbook(nama_file_output)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
            for column in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            item_len = len(str(cell.value))
                            if item_len > max_length:
                                max_length = item_len
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2
            wb.save(nama_file_output)
            print(f"--> SUKSES! Data tersimpan di: {nama_file_output}")
            print("--> --- MEMBERSIHKAN FILE ASLI (XLSX) ---")
            for f in file_sukses:
                try:
                    if os.path.exists(f):
                        os.remove(f)
                        print(f"--> [DELETED] {f}")
                except Exception as e:
                    print(f"--> [GAGAL HAPUS] {f}: {e}")
            print("--> Selesai membersihkan file sumber Excel.")
        except Exception as e:
            print(f"--> FATAL ERROR saat menyimpan Excel: {e}")
            print("--> File asli TIDAK dihapus karena proses gagal.")
    else:
        if semua_file:
            print(
                "--> Gagal proses Excel. Tidak ada data yang berhasil diproses."
            )


def clean_to_text(val):
    if pd.isna(val):
        return ""
    try:
        return str(int(round(float(val))))
    except Exception:
        return str(val)


def clean_to_number(val):
    if pd.isna(val):
        return 0
    try:
        str_val = str(val)
        if str_val.endswith(".0"):
            str_val = str_val[:-2]
        if "." in str_val:
            str_val = str_val.replace(".", "")
        if "," in str_val:
            str_val = str_val.replace(",", ".")
        return int(round(float(str_val)))
    except Exception:
        return 0


def read_file_smart(filepath):
    try:
        return pd.read_excel(filepath, header=3, engine="xlrd", dtype=str)
    except ImportError:
        print(
            "--> [ERROR] Modul 'xlrd' belum terinstall (pip install xlrd)"
        )
        sys.exit()
    except Exception:
        try:
            return pd.read_csv(filepath, header=3, dtype=str)
        except UnicodeDecodeError:
            return pd.read_csv(
                filepath, header=3, encoding="latin1", dtype=str
            )
        except Exception:
            return None


def jalankan_proses_acc():
    try:
        current_folder = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        current_folder = os.getcwd()

    files_to_process = ["Acc.xls"]

    print(f"--> --- MEMULAI PROSES ACC DI: {current_folder} ---")

    cols_map = {
        "Tanggal": "Tanggal",
        "Tgl. Pajak": "Tgl. Pajak",
        "No. Referensi": "No. Referensi",
        "No. Faktur Pajak": "No. Faktur Pajak",
        "Nama Pelanggan": "Nama Pelanggan",
        "Negara Pelanggan": "Negara",
        "Jumlah Pajak": "Jumlah Pajak",
        "Nomor Pajak Pelanggan": "Nomor Pajak",
    }

    text_cols = ["No. Referensi", "No. Faktur Pajak", "Nomor Pajak"]
    number_cols = ["Jumlah Pajak"]

    for filename in files_to_process:
        full_path = os.path.join(current_folder, filename)

        if not os.path.exists(full_path):
            print(f"--> [SKIP] File tidak ditemukan: {filename}")
            continue

        print(f"--> [PROSES] Mengerjakan: {filename}...")

        df = read_file_smart(full_path)

        if df is None:
            print("--> [GAGAL] Format file rusak.")
            continue

        if "Tanggal" in df.columns:
            df = df.dropna(subset=["Tanggal"])

        df = df.rename(columns=cols_map)
        available_cols = [c for c in cols_map.values() if c in df.columns]
        df = df[available_cols].copy()

        for col in text_cols:
            if col in df.columns:
                df[col] = df[col].apply(clean_to_text)

        for col in number_cols:
            if col in df.columns:
                df[col] = df[col].apply(clean_to_number)

        output_filename = "Hasil_CleanerACC.xlsx"
        output_path = os.path.join(current_folder, output_filename)

        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Sheet1")
                worksheet = writer.sheets["Sheet1"]
                for i, column in enumerate(df.columns):
                    try:
                        max_len = (
                            max(
                                df[column].astype(str).map(len).max(),
                                len(column),
                            )
                            + 3
                        )
                    except:
                        max_len = 15
                    col_letter = get_column_letter(i + 1)
                    worksheet.column_dimensions[col_letter].width = max_len

            print(f"--> [SUKSES] Disimpan: {output_filename}")

            try:
                if os.path.exists(full_path):
                    os.remove(full_path)
                    print(f"--> [INFO] File asli dihapus: {filename}")
            except Exception as del_err:
                print(
                    f"--> [WARNING] Gagal menghapus file asli: {del_err}"
                )

        except Exception as e:
            print(
                f"--> [ERROR] Gagal menyimpan (File asli TIDAK dihapus): {e}"
            )

    print("--> --- SELESAI PROSES ACC ---")


if __name__ == "__main__":
    print("--> ========================================")
    print("--> MASTER SCRIPT: AUTO DETECTION MODE")
    print("--> ========================================")

    csv_files = glob.glob("data_export*.csv")
    xlsx_files_export = glob.glob("data_export*.xlsx")
    xlsx_files_export = [f for f in xlsx_files_export if "HMA_" not in f]

    if csv_files or xlsx_files_export:
        print(
            "--> Ditemukan file 'data_export'. Menjalankan Modul Export..."
        )

        if csv_files:
            print(f"--> -> {len(csv_files)} CSV ditemukan.")
            jalankan_proses_csv()

        if xlsx_files_export:
            print(f"--> -> {len(xlsx_files_export)} XLSX ditemukan.")
            jalankan_proses_excel()

    files_acc = ["Acc.xls"]
    found_acc = [f for f in files_acc if os.path.exists(f)]

    if found_acc:
        print(
            f"--> Ditemukan file Acc.xls ({len(found_acc)} file). Menjalankan Modul Acc..."
        )
        jalankan_proses_acc()

    if not (csv_files or xlsx_files_export or found_acc):
        print(
            "--> [INFO] Tidak ditemukan file target apapun (data_export*, Acc.xls)."
        )

    print("--> === SEMUA TUGAS SELESAI ===")