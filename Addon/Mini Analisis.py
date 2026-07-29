import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def buat_laporan_rekonsiliasi_terstruktur(
    file_lama_path="LAMA.xlsx", 
    file_baru_path="BARU.xlsx", 
    output_path="Laporan_Rekonsiliasi_Terstruktur.xlsx"
):
    print("--> Memulai pembuatan laporan rekonsiliasi multi-tab terstruktur...")

    try:
        xls_lama = pd.ExcelFile(file_lama_path)
        xls_baru = pd.ExcelFile(file_baru_path)
    except Exception as e:
        print(f"--> Gagal membaca file Excel. Pastikan file LAMA.xlsx dan BARU.xlsx tersedia. Error: {e}")
        return

    def compare_sheet_multi_stage(df_l, df_b, key_col, col_pajak, col_nama, sheet_name):
        df_l = df_l.copy()
        df_b = df_b.copy()

        df_l[key_col] = df_l[key_col].astype(str).str.strip()
        df_b[key_col] = df_b[key_col].astype(str).str.strip()
        df_l[col_nama] = df_l[col_nama].astype(str).str.strip()
        df_b[col_nama] = df_b[col_nama].astype(str).str.strip()

        matched_l_idx = set()
        matched_b_idx = set()
        diff_details = []

        total_pajak_lama = df_l[col_pajak].sum() if col_pajak in df_l.columns else 0.0
        total_pajak_baru = df_b[col_pajak].sum() if col_pajak in df_b.columns else 0.0

        for idx_b, row_b in df_b.iterrows():
            key_val = row_b[key_col]
            if not key_val or key_val in ['nan', '-']:
                continue
            matches_l = df_l[(df_l[key_col] == key_val) & (~df_l.index.isin(matched_l_idx))]
            if not matches_l.empty:
                idx_l = matches_l.index[0]
                row_l = df_l.loc[idx_l]

                matched_l_idx.add(idx_l)
                matched_b_idx.add(idx_b)

                val_l = float(row_l[col_pajak]) if pd.notna(row_l[col_pajak]) else 0.0
                val_b = float(row_b[col_pajak]) if pd.notna(row_b[col_pajak]) else 0.0
                diff_val = val_b - val_l

                if abs(diff_val) > 0.01:
                    diff_details.append({
                        "Sheet": sheet_name,
                        "Status": "Beda Nominal Pajak",
                        "No. Faktur Pajak": key_val,
                        "Nama Pelanggan": row_b[col_nama],
                        "Kolom Fokus": col_pajak,
                        "Nilai LAMA": val_l,
                        "Nilai BARU": val_b,
                        "Selisih (BARU - LAMA)": diff_val,
                        "Keterangan": f"Nominal {col_pajak} berubah sebesar Rp {diff_val:,.2f}"
                    })

        unmatched_b = df_b[~df_b.index.isin(matched_b_idx)]
        unmatched_l = df_l[~df_l.index.isin(matched_l_idx)].copy()

        for idx_b, row_b in unmatched_b.iterrows():
            nama_b = row_b[col_nama]
            val_b = float(row_b[col_pajak]) if pd.notna(row_b[col_pajak]) else 0.0

            candidates = unmatched_l[
                (unmatched_l[col_nama] == nama_b) & 
                (abs(unmatched_l[col_pajak].astype(float) - val_b) <= 0.01)
            ]

            if not candidates.empty:
                idx_l = candidates.index[0]
                matched_l_idx.add(idx_l)
                matched_b_idx.add(idx_b)
                unmatched_l = unmatched_l.drop(idx_l)

        unmatched_b_s3 = df_b[~df_b.index.isin(matched_b_idx)]
        unmatched_l_s3 = df_l[~df_l.index.isin(matched_l_idx)].copy()

        for idx_b, row_b in unmatched_b_s3.iterrows():
            nama_b = row_b[col_nama]
            val_b = float(row_b[col_pajak]) if pd.notna(row_b[col_pajak]) else 0.0

            candidates = unmatched_l_s3[unmatched_l_s3[col_nama] == nama_b]
            if len(candidates) == 1:
                idx_l = candidates.index[0]
                row_l = unmatched_l_s3.loc[idx_l]
                val_l = float(row_l[col_pajak]) if pd.notna(row_l[col_pajak]) else 0.0
                diff_val = val_b - val_l

                matched_l_idx.add(idx_l)
                matched_b_idx.add(idx_b)
                unmatched_l_s3 = unmatched_l_s3.drop(idx_l)

                if abs(diff_val) > 0.01:
                    diff_details.append({
                        "Sheet": sheet_name,
                        "Status": "Beda Nominal Pajak",
                        "No. Faktur Pajak": f"{row_b[key_col]} (No FP Beda)",
                        "Nama Pelanggan": nama_b,
                        "Kolom Fokus": col_pajak,
                        "Nilai LAMA": val_l,
                        "Nilai BARU": val_b,
                        "Selisih (BARU - LAMA)": diff_val,
                        "Keterangan": f"Revisi/Pembatalan: Nominal selisih Rp {diff_val:,.2f}"
                    })

        unmatched_b_final = df_b[~df_b.index.isin(matched_b_idx)]
        for idx_b, row_b in unmatched_b_final.iterrows():
            val_b = float(row_b[col_pajak]) if pd.notna(row_b[col_pajak]) else 0.0
            diff_details.append({
                "Sheet": sheet_name,
                "Status": "Hanya di BARU (Data Baru)",
                "No. Faktur Pajak": row_b[key_col],
                "Nama Pelanggan": row_b[col_nama],
                "Kolom Fokus": col_pajak,
                "Nilai LAMA": 0.0,
                "Nilai BARU": val_b,
                "Selisih (BARU - LAMA)": val_b,
                "Keterangan": "Transaksi baru ditambahkan di file BARU"
            })

        unmatched_l_final = df_l[~df_l.index.isin(matched_l_idx)]
        for idx_l, row_l in unmatched_l_final.iterrows():
            val_l = float(row_l[col_pajak]) if pd.notna(row_l[col_pajak]) else 0.0
            diff_details.append({
                "Sheet": sheet_name,
                "Status": "Hanya di LAMA (Terhapus)",
                "No. Faktur Pajak": row_l[key_col],
                "Nama Pelanggan": row_l[col_nama],
                "Kolom Fokus": col_pajak,
                "Nilai LAMA": val_l,
                "Nilai BARU": 0.0,
                "Selisih (BARU - LAMA)": -val_l,
                "Keterangan": "Transaksi dihapus / pembatalan di file BARU"
            })

        summary = {
            "Total Nominal LAMA": total_pajak_lama,
            "Total Nominal BARU": total_pajak_baru,
            "Selisih Total Nominal": total_pajak_baru - total_pajak_lama,
            "Total Record LAMA": len(df_l),
            "Total Record BARU": len(df_b),
            "Sama Identik / Matched": len(matched_b_idx) - len([d for d in diff_details if d["Status"] == "Beda Nominal Pajak"]),
            "Hanya di BARU": len([d for d in diff_details if d["Status"] == "Hanya di BARU (Data Baru)"]),
            "Hanya di LAMA": len([d for d in diff_details if d["Status"] == "Hanya di LAMA (Terhapus)"]),
            "Beda Nominal": len([d for d in diff_details if d["Status"] == "Beda Nominal Pajak"]),
            "Total Baris Selisih": len(diff_details)
        }

        return summary, pd.DataFrame(diff_details)

    df_c_l = pd.read_excel(xls_lama, sheet_name="Data Asli Coretax")
    df_c_b = pd.read_excel(xls_baru, sheet_name="Data Asli Coretax")
    sum_c, diff_c = compare_sheet_multi_stage(
        df_c_l, df_c_b,
        key_col="Faktur Pajak/Dokumen Tertentu/Nota Retur/Nota Pembatalan - Nomor",
        col_pajak="VAT",
        col_nama="Name",
        sheet_name="Data Asli Coretax"
    )

    df_a_l = pd.read_excel(xls_lama, sheet_name="Data Asli Accurate")
    df_a_b = pd.read_excel(xls_baru, sheet_name="Data Asli Accurate")
    sum_a, diff_a = compare_sheet_multi_stage(
        df_a_l, df_a_b,
        key_col="No. Faktur Pajak",
        col_pajak="Jumlah Pajak",
        col_nama="Nama Pelanggan",
        sheet_name="Data Asli Accurate"
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_title = Font(name="Segoe UI", size=15, bold=True, color="1F4E78")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="595959")
    font_section = Font(name="Segoe UI", size=11, bold=True, color="1F4E78")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_regular = Font(name="Segoe UI", size=10, color="000000")

    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_sub_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    fill_added = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_removed = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_changed = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    border_thin = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    ws_dash = wb.create_sheet(title="Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True

    ws_dash["A1"] = "DASHBOARD REKONSILIASI PAJAK (CORETAX & ACCURATE)"
    ws_dash["A1"].font = font_title
    ws_dash["A2"] = "Ringkasan Eksekutif Perbandingan File LAMA.xlsx vs BARU.xlsx"
    ws_dash["A2"].font = font_subtitle

    ws_dash["A4"] = "1. RINGKASAN TOTAL NOMINAL PAJAK"
    ws_dash["A4"].font = font_section

    headers_fin = ["Metrik Keuangan", "Data Asli Coretax (VAT)", "Data Asli Accurate (Jumlah Pajak)", "Total Kombinasi"]
    for c_idx, h_text in enumerate(headers_fin, 1):
        cell = ws_dash.cell(row=5, column=c_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fin_rows = [
        ("Total Nominal Pajak (File LAMA)", sum_c["Total Nominal LAMA"], sum_a["Total Nominal LAMA"]),
        ("Total Nominal Pajak (File BARU)", sum_c["Total Nominal BARU"], sum_a["Total Nominal BARU"]),
        ("Selisih Total Nominal (BARU - LAMA)", sum_c["Selisih Total Nominal"], sum_a["Selisih Total Nominal"])
    ]

    for r_idx, (label_f, val_c, val_a) in enumerate(fin_rows, 6):
        c_lbl = ws_dash.cell(row=r_idx, column=1, value=label_f)
        c_vc = ws_dash.cell(row=r_idx, column=2, value=val_c)
        c_va = ws_dash.cell(row=r_idx, column=3, value=val_a)
        c_vt = ws_dash.cell(row=r_idx, column=4, value=f"=B{r_idx}+C{r_idx}")

        is_tot_row = (r_idx == 8)
        r_f = font_bold if is_tot_row else font_regular

        for c in [c_lbl, c_vc, c_va, c_vt]:
            c.font = r_f
            c.border = cell_border
            if c != c_lbl:
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left")
        if is_tot_row:
            for c in [c_lbl, c_vc, c_va, c_vt]:
                c.fill = fill_sub_header

    start_r_rec = 10
    ws_dash.cell(row=start_r_rec, column=1, value="2. REKAPITULASI STATUS DATA & RINCIAN SELISIH").font = font_section

    headers_sum = ["Indikator Rekonsiliasi", "Data Asli Coretax", "Data Asli Accurate", "Total Selisih Kombinasi"]
    for c_idx, h_text in enumerate(headers_sum, 1):
        cell = ws_dash.cell(row=start_r_rec+1, column=c_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_rows = [
        ("Total Record File LAMA", sum_c["Total Record LAMA"], sum_a["Total Record LAMA"]),
        ("Total Record File BARU", sum_c["Total Record BARU"], sum_a["Total Record BARU"]),
        ("Data Cocok Identik (Matched)", sum_c["Sama Identik / Matched"], sum_a["Sama Identik / Matched"]),
        ("Data Baru Ditambahkan (Hanya di BARU)", sum_c["Hanya di BARU"], sum_a["Hanya di BARU"]),
        ("Data Dihapus / Pembatalan (Hanya di LAMA)", sum_c["Hanya di LAMA"], sum_a["Hanya di LAMA"]),
        ("Data Berubah Nominal (Beda Pajak)", sum_c["Beda Nominal"], sum_a["Beda Nominal"]),
        ("TOTAL BARIS RINCIAN SELISIH", sum_c["Total Baris Selisih"], sum_a["Total Baris Selisih"])
    ]

    for offset, (label_m, val_c, val_a) in enumerate(summary_rows, start_r_rec+2):
        c_label = ws_dash.cell(row=offset, column=1, value=label_m)
        c_val_c = ws_dash.cell(row=offset, column=2, value=val_c)
        c_val_a = ws_dash.cell(row=offset, column=3, value=val_a)
        c_tot = ws_dash.cell(row=offset, column=4, value=f"=B{offset}+C{offset}")

        is_tot = (offset == start_r_rec + 8)
        r_f = font_bold if is_tot else font_regular

        for c in [c_label, c_val_c, c_val_a, c_tot]:
            c.font = r_f
            c.border = cell_border
            if c != c_label:
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left")

        if is_tot:
            for c in [c_label, c_val_c, c_val_a, c_tot]:
                c.fill = fill_sub_header

    ws_dash.column_dimensions['A'].width = 44
    ws_dash.column_dimensions['B'].width = 30
    ws_dash.column_dimensions['C'].width = 32
    ws_dash.column_dimensions['D'].width = 28

    def build_detail_sheet(ws, title_text, df_diff):
        ws.views.sheetView[0].showGridLines = True
        ws["A1"] = title_text
        ws["A1"].font = font_title

        headers_det = [
            "No", "Nama Sheet", "Status Perubahan", "No. Faktur Pajak", 
            "Nama Pelanggan", "Kolom Fokus", "Nilai di LAMA", "Nilai di BARU", 
            "Selisih (BARU - LAMA)", "Keterangan"
        ]

        for c_idx, h_text in enumerate(headers_det, 1):
            cell = ws.cell(row=3, column=c_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        curr_r = 4
        for idx, row in df_diff.iterrows():
            ws.cell(row=curr_r, column=1, value=idx+1).alignment = Alignment(horizontal="center")
            ws.cell(row=curr_r, column=2, value=row["Sheet"])

            st_cell = ws.cell(row=curr_r, column=3, value=row["Status"])
            if "Hanya di BARU" in row["Status"]:
                st_cell.fill = fill_added
            elif "Hanya di LAMA" in row["Status"]:
                st_cell.fill = fill_removed
            else:
                st_cell.fill = fill_changed

            ws.cell(row=curr_r, column=4, value=row["No. Faktur Pajak"])
            ws.cell(row=curr_r, column=5, value=row["Nama Pelanggan"])
            ws.cell(row=curr_r, column=6, value=row["Kolom Fokus"])

            ws.cell(row=curr_r, column=7, value=row["Nilai LAMA"])
            ws.cell(row=curr_r, column=8, value=row["Nilai BARU"])
            ws.cell(row=curr_r, column=9, value=row["Selisih (BARU - LAMA)"])
            ws.cell(row=curr_r, column=10, value=row["Keterangan"])

            for c_i in range(1, 11):
                cell = ws.cell(row=curr_r, column=c_i)
                cell.font = font_regular
                cell.border = cell_border
                if c_i in [7, 8, 9]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

            curr_r += 1

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 26
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 22
        ws.column_dimensions['J'].width = 45

    ws_c = wb.create_sheet(title="Rincian Coretax")
    build_detail_sheet(ws_c, "RINCIAN SELISIH DATA ASLI CORETAX (VAT)", diff_c)

    ws_a = wb.create_sheet(title="Rincian Accurate")
    build_detail_sheet(ws_a, "RINCIAN SELISIH DATA ASLI ACCURATE (JUMLAH PAJAK)", diff_a)

    ws_kon = wb.create_sheet(title="Rincian Konsolidasi")
    df_combined = pd.concat([diff_c, diff_a], ignore_index=True)
    build_detail_sheet(ws_kon, "RINCIAN KONSOLIDASI SELISIH CORETAX & ACCURATE", df_combined)

    wb.save(output_path)
    print(f"--> Laporan berhasil dibuat dan disimpan di: {output_path}")

if __name__ == "__main__":
    buat_laporan_rekonsiliasi_terstruktur("LAMA.xlsx", "BARU.xlsx", "Laporan_Rekonsiliasi.xlsx")